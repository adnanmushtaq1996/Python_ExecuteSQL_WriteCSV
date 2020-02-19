#Analysis -CSV and XLSX

#Import all necessary dependancies
import pyodbc,csv
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font,Border,PatternFill
import os


#Current date UTC
now = datetime.utcnow()

#Days to go back
print("Enter the Duration in days from the current date you want to show records for!!");
z = int(input())
print("Printing Records ",z,"Days before current Date");
y = now - timedelta(days=z)

print("Current System Date : ",now);
print("Date to go back to : ",y);

#make  connection with the database server
conn = pyodbc.connect(
"Driver={ODBC Driver 17 for SQL Server};"
"Server=;" # enter Server Name
"Database=;" # enter Database Name
"Trusted_Connection=yes;"
)

c = conn.cursor()

#query 1
c.execute("""SELECT * FROM DATABASENAME  
WHERE DATE BETWEEN ? AND ?""",y,now )  # Enter the query to be executed

res1=c.fetchall()
#Print query output to Console
for row1 in res1:
    print(f" ColumnName1={row1[0]} ColumnName2 ={row1[1]} ColumnName3={row1[2]} ColumnName4={row1[3]}")
    
#query 2    
c.execute("""SELECT * FROM DATABASENAME  
WHERE DATE BETWEEN ? AND ?""",y,now ) # Enter the query to be executed


res2=c.fetchall()
#Print query output to Console
for row2 in res2:
    print(f" ColumnName2={row2[0]} ColumnName5 ={row2[1]}  ")

#query 3    
c.execute("""SELECT * FROM DATABASENAME  
WHERE DATE BETWEEN ? AND ?""",y,now )  # Enter the query to be executed

res3=c.fetchall()

#Print query output to Console
for row3 in res3:
    print(f"  ColumnName2 ={row3[0]} ColumnName6={row3[1]} ")
    
#query 4    
c = conn.cursor()

sql= c.execute("""SELECT * FROM DATABASENAME  
WHERE DATE BETWEEN ? AND ?""",y,now)  # Enter the query to be executed
res4=c.fetchall()

#Print query output to Console
for row4 in res4:
    print(f"  ColumnName2 ={row4[0]} ColumnName7={row3[1]} ")

#Generating intermediate csv files(optional)
header = ['ColumnName1', 'ColumnName2', 'ColumnName3','ColumnName4']
with open('temp1.csv', 'wt', newline ='') as file:
    writer = csv.writer(file, delimiter=',')
    writer.writerow(i for i in header)
    for j in res1:
        writer.writerow(j)
file.close()

header = [ 'ColumnName2', 'ColumnName5']
with open('temp2.csv', 'wt', newline ='') as file:
    writer = csv.writer(file, delimiter=',')
    writer.writerow(i for i in header)
    for j in res2:
        writer.writerow(j)
file.close()

header = ['ColumnName2','ColumnName6']
with open('temp3.csv', 'wt', newline ='') as file:
    writer = csv.writer(file, delimiter=',')
    writer.writerow(i for i in header)
    for j in res3:
        writer.writerow(j)
file.close()

header = ['ColumnName2', 'ColumnName7']
with open('temp4.csv', 'wt', newline ='') as file:
    writer = csv.writer(file, delimiter=',')
    writer.writerow(i for i in header)
    for j in res4:
        writer.writerow(j)
file.close()


#Pandas for generating the final csv
df1 = pd.read_csv("temp1.csv")
df2 = pd.read_csv("temp2.csv")
df3 = pd.read_csv("temp3.csv")
df4 = pd.read_csv("temp4.csv")


#Combine on basis of ColumnName2
df5=df1.merge(df2, on="ColumnName2")
df6=df3.merge(df4, on="ColumnName2")

#Remove Intermediate excels(optional)
os.remove('temp1.csv')
os.remove('temp2.csv')
os.remove('temp3.csv')
os.remove('temp4.csv')


#FINAL CSV
df=df5.merge(df6, on="ColumnName2")
df['ColumnName8_Unrounded']=df['ColumnName3'] /df['ColumnName4']  #If you want to perform some math on columns
df['ColumnName8']=df['ColumnName8_Unrounded'].round(decimals=2)   #If you want to round off some column value
del df['Hours Produced per HIL_Unrounded']
print(df)
df.to_csv(r'Final_CSV.csv')

#Generate FINAL XLSX
wb = Workbook()
ws = wb.active
with open('Final_CSV.csv', 'r') as f:
    for row in csv.reader(f):
        ws.append(row)

    # Beautify Header and Excel(optional)
    Bold_header=Font(bold=True,color='fcc133')
    ws['B1'].font=ws['C1'].font=ws['D1'].font=ws['E1'].font=ws['F1'].font=ws['G1'].font=ws['H1'].font= ws['I1'].font=Bold_header
    colorFill = PatternFill(start_color='292930',end_color='292930', fill_type='solid')
    ws['B1'].fill =ws['C1'].fill=ws['D1'].fill=ws['E1'].fill=ws['F1'].fill=ws['G1'].fill=ws['H1'].fill= ws['I1'].fill= colorFill


wb.save('Final_EXCEL.xlsx')






